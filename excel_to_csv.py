# Excel to CSV converter
from pathlib import Path
import openpyxl, csv, sys
print('=' * 40)
print('CSV Converter')
print('=' * 40)
while True:
    excel_file_path = Path(input('Enter Excel file path: ').strip())
    if excel_file_path.is_file():
        break
    else:
        print('Invalid file path! Please try again.\n')
# make new folder and csv file
folder_parent = excel_file_path.parent
file_name = excel_file_path.stem
csv_folder = Path(folder_parent) / f'converted'
csv_folder.mkdir( parents = True, exist_ok = True)
converted_file = csv_folder / f'{file_name}.csv'
# read excel file
try:
    wb = openpyxl.load_workbook(excel_file_path)
except Exception as e:
    sys.exit(f'Couldn\'t open workbook.\n{e}')
while True:
    user_sheet = input(f'\nAvailable sheets:{wb.sheetnames}\nChoose worksheet: (case sensitive)').strip()
    if user_sheet in wb.sheetnames:
        break
    print('Invalid sheet name!')
sheet = wb[user_sheet]
row_counts = 0
# open new CSV file and write
with open(converted_file, 'w', encoding = 'utf-8', newline = '') as file:
    writer = csv.writer(file)
    for row in sheet.iter_rows(values_only = True):
        writer.writerow(row)
        row_counts += 1
wb.close()
print('-------------------\n Conversion completed.\n-------------------')
print(f'\nWorkbook: {file_name}.xlsx')
print(f'\nWorksheet: \n{user_sheet}')
print(f'\nRows exported: \n{row_counts}')
print(f'\nCSV saved at: \n{converted_file}')